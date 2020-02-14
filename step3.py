import openpyxl
from openpyxl import load_workbook


def open_excel(excel_data_url):
    #excel_sheet = openpyxl.load_workbook(excel_data_url)
    global endrow
    wb = load_workbook(excel_data_url)
    sh = wb["Sheet1"]
    #print(excel_sheet.sheetnames)
    #excel_data = excel_sheet['sheet1']
    #print(len(excel_data.rows))
    num = 0
    start = []
    error_index = []
    endrow = 0
    flag = 0
    final_list = []
    # print(sh.max_row)
    for row in range(2, sh.max_row):
        #for col in range(1, 14):

            if sh.cell(row=row, column=2).value == 0 and sh.cell(row=row, column=10).value != 0:
                #if flag == 1:
                    error_index.append(row)
                    num += 1  # 共有多少个异常数据
                #print(row)
                # 获取某行某列单元格元素
                # start.append(row)

                # if num < 180:
                #     flag = 1
                # elif num == 180:
                #     flag = 2
                # else:
                #     flag = 3
                # endrow = row    # 取最后一个异常的行数
            # elif sh.cell(row=row, column=2).value < 10 and num > 180:
            #     # print('1111')
            #     error_index.append(row)
            #     num = num + 1
            if sh.cell(row=row, column=2).value != 0 and num <= 180:
                error_index = []
                num = 0
                continue
            if sh.cell(row=row, column=2).value != 0 and num > 180:
                final_list += error_index
                num = 0

            #print(sh.cell(row=row, column=col).value, end=',')
        #print()

    # for i in range(2, sh.max_row):
    #     if i in final_list:
    # 输出行号长度
    print(len(final_list))
    count = 0
    # print(final_list[0])
    start_end = []
    start_end.append(final_list[0])
    # print(final_list[0]+1 == final_list[1])
    for i in range(1, len(final_list)-2):
        # count += 1
        # print(count)
        # sh.delete_rows(final_list[i])
            if final_list[i-1]+1 == final_list[i]:
                pass
            elif final_list[i-1]+1 != final_list[i]:
                start_end.append(final_list[i-1])
                start_end.append(final_list[i])
    start_end.append(final_list[len(final_list)-1])
    for i in range(0, len(start_end)):
        if i % 2 == 0:
            print('start')
            print(start_end[i])
        else:
            print('end')
            print(start_end[i])
        # print(final_list[i])
    # wb.save(excel_data_url)



    # startrow = start[0]
    #     print(startrow)
    #     print(endrow)
    # print(num)

    # 删除GPS车速为0的数据（大于180条）
    # count = 0
    # for row in range(startrow, endrow+1):
    #     count = count + 1
    #     print(count)
    #     sh.delete_rows(startrow)
    # wb.save(excel_data_url)


def main():
    excel_data_url = 'step3_data3.xlsx'
    open_excel(excel_data_url)


if __name__ == '__main__':
    main()