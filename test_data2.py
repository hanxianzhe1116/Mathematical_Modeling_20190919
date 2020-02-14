from openpyxl import load_workbook

'''
    打开excel表，读取数据并进行筛选
'''
def open_excel(excel_data_url):

    wb = load_workbook(excel_data_url)
    sh = wb["Sheet1"]

    '''
        在excel表中找出异常数据
        异常数据定义：
            1.GPS车速为0，
            2.瞬时油耗不为0，
            3.以上2种情况连续出现超过180秒
    '''
    num = 0  # 记录连续异常数据的数量
    error_index = []  # 暂存异常数据
    final_list = []  # 存储最终异常数据的索引
    for row in range(2, sh.max_row):
        if sh.cell(row=row, column=2).value == 0 and sh.cell(row=row, column=10).value != 0:
            error_index.append(row)
            num += 1  # 共有多少个异常数据
        if sh.cell(row=row, column=2).value != 0 and num <= 180:
            error_index = []
            num = 0
            continue
        if sh.cell(row=row, column=2).value != 0 and num > 180:
            final_list += error_index
            num = 0
    print(len(final_list))

    '''
        找出异常数据的起始索引
    '''
    start_end = []
    start_end.append(final_list[0])
    for i in range(1, len(final_list)-2):
        if final_list[i-1]+1 == final_list[i]:
            pass
        elif final_list[i-1]+1 != final_list[i]:
            start_end.append(final_list[i-1])
            start_end.append(final_list[i])
    start_end.append(final_list[len(final_list)-1])

    '''
        输出异常数据的起始索引
    '''
    for i in range(0, len(start_end)):
        if i % 2 == 0:
            print('第'+str(i+1)+'组 :')
            print('start')
            print(start_end[i])
        else:
            print('end')
            print(start_end[i])

    '''
        在excel中删除找出的异常数据
    '''
    count = 0
    for i in final_list:
        count += 1
        print(count)
        sh.delete_rows(final_list[i])
        # print(final_list[i])
    wb.save(excel_data_url)


def main():
    excel_data_url = 'step3_data2.xlsx'  # 数据表
    open_excel(excel_data_url)


if __name__ == '__main__':
    main()