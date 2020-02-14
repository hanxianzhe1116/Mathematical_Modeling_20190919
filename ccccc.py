# ----coding:utf-8
# ----create-time:2019/9/19
from openpyxl import *
import pandas as pd
from dateutil.parser import parse
import datetime as datetime

# wb = load_workbook('test.xlsx')
wb = load_workbook('original_data3.xlsx')
ws=wb['Sheet1']
data01=[]
for i in range(2,ws.max_row,1):
    data01.append(ws.cell(row=i,column=1).value)
data01=pd.DataFrame(data01)

car_time = (data01.iloc[:, 0])
car_time=car_time.apply(lambda x:x[:-5])
car_time=car_time.apply(lambda x:parse(x))
break_time = []
for index in range(0,len(car_time)-1,1):
      if (car_time[index+1]-car_time[index])!=datetime.timedelta(seconds=1):
         # if(car_time[index+1]-car_time[index]>datetime.timedelta(seconds=900)):
             break_time.append({'before':car_time[index],'after':car_time[index+1]})


# wb1 = wb.active  #激活sheet
# wb1.cell(2,2,'pass2')#往sheet中的第二行第二列写入‘pass2’的数据
# wb.save('test.xlsx')#保存
print(int(len(break_time)/2))
print(break_time)
