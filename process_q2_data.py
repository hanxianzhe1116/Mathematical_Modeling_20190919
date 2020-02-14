from openpyxl import load_workbook

'''
    打开excel表，读取数据并进行筛选
'''
def open_excel(excel_data_url):
    wb = load_workbook(excel_data_url)
    sh = wb["Sheet1"]

    '''
        在excel表中找出运动学片段
        经过数据筛选后，数据表现正常，根据运动学片段定义：
            从怠速状态到下一个怠速状态的运动片段称为运动学片段
            怠速： 速度为 0 ，发动机转速不为 0
        遍历GPS车速
    '''
    num = 0  # 记录运动学片段数量
    flag = 0
    for row in range(2, sh.max_row):
        if sh.cell(row=row, column=2).value == 0 and sh.cell(row=row+1, column=2).value == 0:
            flag = 1
            continue
        else:
            flag = 0
        if sh.cell(row=row, column=2).value != 0 and sh.cell(row=row+1, column=2).value != 0:
            continue
        if flag == 0:
            num += 1
    print('预处理文件3中共有'+str(int(num/2))+'组运动学片段')


def main():
    excel_data_url = 'step3_data3_q2.xlsx'  # 数据表
    open_excel(excel_data_url)


if __name__ == '__main__':
    main()
