

from openpyxl import *
import pandas as pd
from dateutil.parser import parse
import datetime as datetime

filename = 'data____1.xlsx'
wb = load_workbook(filename)
wss = wb.active
ws = wb['sheet1']
data01 = []

for i in range(2, ws.max_row+1, 1):
    data01.append(ws.cell(row=i, column=1).value)  # 获取单列的数据，存入列表中
data01 = pd.DataFrame(data01)
car_time = (data01.iloc[:, 0])
car_time=car_time.apply(lambda x: x[:-5])  # 处理时间后五位为0的情况
car_time=car_time.apply(lambda x: parse(x))
break_time = []
print(car_time)

'''
    查找时间断点，并插入行，补充数据
'''
for index in range(0, len(car_time)-1):
    if (car_time[index + 1] - car_time[index]) != datetime.timedelta(seconds=1):
        # if(car_time[index+1]-car_time[index]>datetime.timedelta(seconds=900)):
        break_time.append(
            {'index': index, 'before': car_time[index], 'index+1': index + 1, 'after': car_time[index + 1]})
        ws.insert_rows(index+len(break_time)+2)  # 插入行数
        print(car_time[index + 1] - car_time[index])
        for second in [car_time[index+1]-car_time[index]]:
            ws.cell(index+len(break_time)+2, 1, str(car_time[index]+datetime.timedelta(seconds=second))+'.000.')
wb.save(filename)
wb1 = wb.active  # 激活sheet
# wb1.cell(2, 2, 'pass2')#往sheet中的第二行第二列写入‘pass2’的数据
wb1.save('processed_data1.xlsx')  # 保存至新的excel表
print(len(break_time))
print(break_time)
